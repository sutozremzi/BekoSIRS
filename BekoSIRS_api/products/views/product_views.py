# products/views/product_views.py
"""
Product and Category management views.
"""

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from django.db.models import Count
from django.http import HttpResponse

from products.models import Product, Category, ProductOwnership, WishlistItem, Notification
from products.serializers import ProductSerializer, CategorySerializer


class ProductViewSet(viewsets.ModelViewSet):
    """Product CRUD operations with role-based access."""
    queryset = Product.objects.all().select_related("category")
    serializer_class = ProductSerializer

    def get_permissions(self):
        if self.action in ["list", "retrieve"]:
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_queryset(self):
        return Product.objects.all().select_related("category")

    @action(
        detail=False,
        methods=["get"],
        url_path="my-products",
        permission_classes=[IsAuthenticated],
    )
    def my_products(self, request):
        """GET /api/products/my-products/ - User's assigned products."""
        user = request.user

        if user.role in ["admin", "seller"]:
            qs = Product.objects.all().select_related("category")
            return Response(ProductSerializer(qs, many=True).data)

        ownerships = (
            ProductOwnership.objects.filter(customer=user)
            .select_related("product", "product__category")
            .order_by("-id")
        )

        result = []
        for o in ownerships:
            p = o.product
            item = ProductSerializer(p).data
            if hasattr(o, "assigned_date"):
                item["assigned_date"] = o.assigned_date
            elif hasattr(o, "assigned_at"):
                item["assigned_date"] = o.assigned_at
            elif hasattr(o, "created_at"):
                item["assigned_date"] = o.created_at
            if hasattr(o, "status"):
                item["status"] = o.status
            result.append(item)

        return Response(result)

    def perform_update(self, serializer):
        """Detect price changes and send notifications."""
        instance = self.get_object()
        old_price = instance.price
        new_price = serializer.validated_data.get('price', old_price)
        updated_instance = serializer.save()

        if new_price and new_price < old_price:
            self._send_price_drop_notifications(updated_instance, old_price, new_price)

    def _send_price_drop_notifications(self, product, old_price, new_price):
        """Send price drop notifications to wishlist users."""
        discount_percent = round((float(old_price) - float(new_price)) / float(old_price) * 100, 1)
        
        wishlist_items = WishlistItem.objects.filter(
            product=product,
            notify_on_price_drop=True
        ).select_related('wishlist__customer')

        notifications = []
        for item in wishlist_items:
            user = item.wishlist.customer
            if user.notify_price_drops:
                notifications.append(
                    Notification(
                        user=user,
                        notification_type='price_drop',
                        title=f'Fiyat Düştü! %{discount_percent} İndirim',
                        message=f'{product.name} ürününün fiyatı {old_price}₺ yerine {new_price}₺ oldu!',
                        related_product=product
                    )
                )

        if notifications:
            Notification.objects.bulk_create(notifications)

    @action(
        detail=False,
        methods=['post'],
        url_path='compare',
        permission_classes=[IsAuthenticated]
    )
    def compare(self, request):
        """
        POST /api/products/compare/
        Compare multiple products side by side.
        
        Request body:
            {"product_ids": [1, 2, 3]}
        
        Returns:
            - products: List of product details
            - comparison_fields: Fields being compared
            - differences: Highlighted differences
            - recommendation: Best value recommendation
        """
        product_ids = request.data.get('product_ids', [])
        
        # Validation
        if not product_ids:
            return Response(
                {'error': 'product_ids listesi gerekli'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(product_ids) < 2:
            return Response(
                {'error': 'En az 2 ürün seçmelisiniz'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(product_ids) > 4:
            return Response(
                {'error': 'En fazla 4 ürün karşılaştırılabilir'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Fetch products
        products = Product.objects.filter(id__in=product_ids).select_related('category')
        
        if products.count() != len(product_ids):
            return Response(
                {'error': 'Bir veya daha fazla ürün bulunamadı'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Serialize products
        products_data = ProductSerializer(products, many=True).data
        
        # Calculate differences
        comparison_fields = [
            {'key': 'price', 'label': 'Fiyat', 'unit': '₺', 'type': 'currency'},
            {'key': 'brand', 'label': 'Marka', 'type': 'text'},
            {'key': 'warranty_duration_months', 'label': 'Garanti Süresi', 'unit': 'ay', 'type': 'number'},
            {'key': 'stock', 'label': 'Stok', 'unit': 'adet', 'type': 'number'},
            {'key': 'category_name', 'label': 'Kategori', 'type': 'text'},
        ]
        
        # Add category_name to product data
        for pd, p in zip(products_data, products):
            pd['category_name'] = p.category.name if p.category else '-'
        
        # Find differences and highlights
        differences = {}
        for field in comparison_fields:
            key = field['key']
            values = [p.get(key) for p in products_data]
            
            # Check if all values are the same
            unique_values = set(str(v) for v in values)
            is_different = len(unique_values) > 1
            
            # Find best value for numeric fields
            best_indices = []
            if is_different and field.get('type') in ['currency', 'number']:
                numeric_values = []
                for v in values:
                    try:
                        numeric_values.append(float(v) if v else 0)
                    except (TypeError, ValueError):
                        numeric_values.append(0)
                
                if key == 'price':
                    # Lower price is better
                    min_val = min(numeric_values)
                    best_indices = [i for i, v in enumerate(numeric_values) if v == min_val]
                else:
                    # Higher is better (warranty, stock)
                    max_val = max(numeric_values)
                    best_indices = [i for i, v in enumerate(numeric_values) if v == max_val]
            
            differences[key] = {
                'is_different': is_different,
                'values': values,
                'best_indices': best_indices
            }
        
        # Calculate overall recommendation (best value)
        scores = [0] * len(products_data)
        for key, diff in differences.items():
            for idx in diff.get('best_indices', []):
                scores[idx] += 1
        
        best_score = max(scores)
        recommended_indices = [i for i, s in enumerate(scores) if s == best_score]
        
        return Response({
            'products': products_data,
            'comparison_fields': comparison_fields,
            'differences': differences,
            'recommendation': {
                'indices': recommended_indices,
                'product_ids': [products_data[i]['id'] for i in recommended_indices],
                'reason': 'En iyi fiyat/performans oranı'
            }
        })


class CategoryViewSet(viewsets.ModelViewSet):
    """Category CRUD with product count annotation."""
    queryset = Category.objects.annotate(product_count=Count('products')).all()
    serializer_class = CategorySerializer

    def get_permissions(self):
        if self.action in ["list", "retrieve"]:
            return [AllowAny()]
        return [IsAdminUser()]


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def my_products_direct(request):
    """GET /api/my-products/ - Direct endpoint for mobile compatibility."""
    user = request.user

    if user.role in ["admin", "seller"]:
        qs = Product.objects.all().select_related("category")
        return Response(ProductSerializer(qs, many=True).data)

    ownerships = (
        ProductOwnership.objects.filter(customer=user)
        .select_related("product", "product__category")
        .order_by("-id")
    )

    result = []
    for o in ownerships:
        p = o.product
        item = ProductSerializer(p).data
        if hasattr(o, "assigned_date"):
            item["assigned_date"] = o.assigned_date
        elif hasattr(o, "assigned_at"):
            item["assigned_date"] = o.assigned_at
        elif hasattr(o, "created_at"):
            item["assigned_date"] = o.created_at
        if hasattr(o, "status"):
            item["status"] = o.status
        result.append(item)

    return Response(result)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_products_excel(request):
    """GET /api/products/export/excel/ - Export products as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    user = request.user
    if user.role not in ['admin', 'seller']:
        return Response({'error': 'Yetkisiz erişim'}, status=status.HTTP_403_FORBIDDEN)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ürünler"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        "ID", "Ürün Adı", "Marka", "Model Kodu", "Kategori",
        "List Fiyatı (₺)", "Peşin Fiyatı (₺)", "Stok", 
        "Garanti (Ay)", "Garanti Kodu", "Kampanya"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    products = Product.objects.all().select_related('category')

    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.id).border = thin_border
        ws.cell(row=row, column=2, value=product.name).border = thin_border
        ws.cell(row=row, column=3, value=product.brand).border = thin_border
        ws.cell(row=row, column=4, value=getattr(product, 'model_code', '')).border = thin_border
        ws.cell(row=row, column=5, value=product.category.name if product.category else '').border = thin_border
        ws.cell(row=row, column=6, value=float(product.price) if product.price else 0).border = thin_border
        ws.cell(row=row, column=7, value=float(getattr(product, 'price_cash', 0) or 0)).border = thin_border
        ws.cell(row=row, column=8, value=product.stock).border = thin_border
        ws.cell(row=row, column=9, value=product.warranty_duration_months).border = thin_border
        ws.cell(row=row, column=10, value=getattr(product, 'warranty_code', '')).border = thin_border
        ws.cell(row=row, column=11, value=getattr(product, 'campaign_tag', '')).border = thin_border

    column_widths = [8, 40, 15, 15, 20, 15, 15, 10, 12, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bekosirs_products.xlsx"'
    wb.save(response)
    return response
